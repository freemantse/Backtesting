"""Convert the hand-delivered LSEG (Refinitiv) `.xlsx` workbooks to clean per-series CSVs.

The LSEG workbooks in ``data/China/`` are NOT flat tables — each leads with a metadata
header, a Volume-At-Price distribution, and a statistics block, then the real history
table starts ~30 rows down and is ordered newest-first. A naive "Save As CSV" would dump
all that preamble and break the loader. This script extracts only the data region:

Each CSV keeps the workbook's basename (e.g. ``Gold_518880.xlsx`` -> ``Gold_518880.csv``)
and is written alongside the source in ``data/China/`` for easy identification:

  * Price/index files ("Price History" layout) -> ``<Name_Code>.csv`` (date,close):
      scan for the ``Exchange Date`` header row (not a fixed skiprows), read Exchange Date
      + Close, convert any Excel date serials, drop blanks, sort ascending.
  * NBS PMI ("First Release Data" sheet) -> ``PMI_NBS.csv``
      (release_date,period,first_release): a true first-release/point-in-time series, so the
      Original Release Date is preserved as the effective (PIT) date.

This is a ONE-TIME conversion tool — openpyxl is only needed here, not at backtest time
(the run path reads the CSVs via ``--source csvdir``). It fetches/fabricates nothing; a
missing workbook is logged and skipped (never zero-filled — the repo's graceful-degradation
rule). Re-run it whenever LSEG re-delivers an updated workbook (it overwrites the CSV in place).

Usage:  python scripts/convert_lseg_xlsx.py [--src data/China] [--out data/China]
"""
from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent

# "Price History" workbooks (the leg's series code is shown in the comment).
PRICE_FILES = [
    "Growth_159915.xlsx",      # E Fund ChiNext ETF (code 159915)
    "Gold_518880.xlsx",        # Hua An YiFu Gold ETF (code 518880)
    "FreeCashFlow_CNI.xlsx",   # CNI Free Cash Flow index (code 980092)
    "DividendLowVol_CSI.xlsx",  # CSI Dividend Low-Vol index (code H30269)
]
PMI_FILE = "PMI_NBS.xlsx"

# Wind Financial Terminal exports are FLAT tables (header on row 1, newest-first,
# sometimes a trailing "Source: Wind" footer; commodity values carry thousands commas).
# Matched by a substring of the (dated) filename so re-pulls still resolve.
WIND_SPECS = [
    {"match": "CBA21801", "out": "Bond_ChinaBond30Y.csv", "close": "Close"},          # ChinaBond 30Y TR
    {"match": "IMCI", "out": "Commodity_SHFENonferrous.csv", "close": "Closing Price"},  # SHFE non-ferrous
]

EXCEL_EPOCH = pd.Timestamp("1899-12-30")  # Excel serial-date origin


def _to_date(v) -> pd.Timestamp | None:
    """Coerce a cell to a date — handles datetime, Excel serial, or text."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return pd.Timestamp(v).normalize()
    if isinstance(v, (int, float)):  # Excel serial date
        return (EXCEL_EPOCH + pd.Timedelta(days=float(v))).normalize()
    return pd.to_datetime(str(v), errors="coerce")


def convert_price(path: Path) -> pd.DataFrame:
    """Extract the Exchange Date + Close history from a 'Price History' workbook."""
    ws = load_workbook(path, read_only=True, data_only=True).worksheets[0]
    rows = list(ws.iter_rows(values_only=True))

    # Scan for the history header row: the one whose cells include 'Exchange Date' + 'Close'.
    header_idx = date_col = close_col = None
    for i, row in enumerate(rows):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if "Exchange Date" in cells and "Close" in cells:
            header_idx = i
            date_col = cells.index("Exchange Date")
            close_col = cells.index("Close")
            break
    if header_idx is None:
        raise ValueError(f"{path.name}: no 'Exchange Date'/'Close' header row found")

    recs = []
    for row in rows[header_idx + 1:]:
        d = _to_date(row[date_col])
        c = pd.to_numeric(row[close_col], errors="coerce")
        if d is None or pd.isna(d) or pd.isna(c):
            continue  # blank / non-trading / footer row — never zero-fill
        recs.append((d, float(c)))

    df = pd.DataFrame(recs, columns=["date", "close"])
    df = df.drop_duplicates(subset="date").sort_values("date").reset_index(drop=True)
    return df


def convert_pmi(path: Path) -> pd.DataFrame:
    """Extract Period | Original Release Date | First Release from the PMI workbook."""
    wb = load_workbook(path, read_only=True, data_only=True)
    if "First Release Data" not in wb.sheetnames:
        raise ValueError(f"{path.name}: no 'First Release Data' sheet")
    rows = list(wb["First Release Data"].iter_rows(values_only=True))

    # Header row: first cell 'Period', with 'Original Release Date' + 'First Release'.
    header_idx = None
    for i, row in enumerate(rows):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if "Original Release Date" in cells and "First Release" in cells:
            header_idx = i
            break
    if header_idx is None:
        raise ValueError(f"{path.name}: no PMI header row found")

    recs = []
    for row in rows[header_idx + 1:]:
        period, rel, val = (row + (None, None, None))[:3]
        v = pd.to_numeric(val, errors="coerce")
        rel_dt = pd.to_datetime(str(rel), errors="coerce") if rel else None
        per_dt = pd.to_datetime(str(period), errors="coerce")  # "May 2026" -> month start
        if pd.isna(v) or rel_dt is None or pd.isna(rel_dt):
            continue
        if not pd.isna(per_dt):
            per_dt = per_dt + pd.offsets.MonthEnd(0)  # the month the reading refers to
        recs.append((rel_dt.normalize(), per_dt, float(v)))

    df = pd.DataFrame(recs, columns=["release_date", "period", "first_release"])
    df = df.drop_duplicates(subset="release_date").sort_values("release_date").reset_index(drop=True)
    return df


def convert_wind(path: Path, close_header: str) -> pd.DataFrame:
    """Extract Trading Date + <close_header> from a flat Wind 'History Price' export."""
    ws = load_workbook(path, read_only=True, data_only=True).worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"{path.name}: empty sheet")
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    if "Trading Date" not in header or close_header not in header:
        raise ValueError(
            f"{path.name}: missing 'Trading Date'/'{close_header}' header; got {header[:6]}"
        )
    dcol, ccol = header.index("Trading Date"), header.index(close_header)

    recs = []
    for row in rows[1:]:
        d = _to_date(row[dcol])
        raw = row[ccol]
        if isinstance(raw, str):
            raw = raw.replace(",", "").strip()  # commodity values like "4,610.69"
        c = pd.to_numeric(raw, errors="coerce")
        if d is None or pd.isna(d) or pd.isna(c):
            continue  # 'Source: Wind' footer / blank — never zero-fill
        recs.append((d, float(c)))

    df = pd.DataFrame(recs, columns=["date", "close"])
    return df.drop_duplicates(subset="date").sort_values("date").reset_index(drop=True)


def _coverage(df: pd.DataFrame, date_col: str) -> str:
    dates = pd.to_datetime(df[date_col])
    gaps = int((dates.diff().dt.days.fillna(0) > 7).sum())  # holes > 1 week
    return (f"{len(df):>5} rows  {dates.min().date()} -> {dates.max().date()}  "
            f"({gaps} gap(s) >7d)")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--src", default=str(ROOT / "data" / "China"))
    ap.add_argument("--out", default=str(ROOT / "data" / "China"))
    args = ap.parse_args()

    src, out = Path(args.src), Path(args.out)
    out.mkdir(parents=True, exist_ok=True)
    written = []

    for fname in PRICE_FILES:
        path = src / fname
        if not path.exists():
            print(f"[skip] {fname}: not delivered yet (leg will be skipped at run time)")
            continue
        df = convert_price(path)
        if df.empty:
            print(f"[warn] {fname}: no data rows extracted — skipping")
            continue
        dst = out / f"{path.stem}.csv"
        df.to_csv(dst, index=False)
        print(f"[ok]   {fname:<26} -> {dst.name:<24} {_coverage(df, 'date')}")
        written.append(dst)

    pmi_path = src / PMI_FILE
    if pmi_path.exists():
        df = convert_pmi(pmi_path)
        if not df.empty:
            dst = out / f"{pmi_path.stem}.csv"
            df.to_csv(dst, index=False)
            print(f"[ok]   {PMI_FILE:<26} -> {dst.name:<24} {_coverage(df, 'release_date')}")
            written.append(dst)
        else:
            print(f"[warn] {PMI_FILE}: no PMI rows extracted")
    else:
        print(f"[skip] {PMI_FILE}: not delivered yet")

    # Wind exports (Bond, Commodity) — flat tables; remove the source xlsx after
    # a clean write so data/China/ stays CSV-only (same convention as the LSEG files).
    for spec in WIND_SPECS:
        matches = sorted(p for p in src.glob("*.xlsx") if spec["match"] in p.name)
        if not matches:
            print(f"[skip] Wind {spec['match']}: no .xlsx found in {src}")
            continue
        path = matches[-1]
        df = convert_wind(path, spec["close"])
        if df.empty:
            print(f"[warn] {path.name}: no data rows extracted — skipping")
            continue
        dst = out / spec["out"]
        df.to_csv(dst, index=False)
        print(f"[ok]   {path.name:<42} -> {dst.name:<28} {_coverage(df, 'date')}")
        written.append(dst)
        if dst.exists() and dst.stat().st_size > 0:
            path.unlink()
            print(f"       removed source {path.name} (CSV-only)")

    if not written:
        print("[error] no workbooks converted — is the source directory correct?",
              file=sys.stderr)
        return 1
    print(f"\n[done] wrote {len(written)} CSV(s) to {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
